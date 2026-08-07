"""
ISO 286 EdTech - Backend FastAPI (Layered Architecture + JWT Auth/RBAC)

Kiến trúc 3 lớp (Layered Architecture):
- Presentation Layer : app.api        -> router /api/export/docx, /api/export/pdf, /api/iso/tables
- Business Logic     : app.services   -> ⭐ LỚP NGHIỆP VỤ (BLL)
- Configuration      : app.core       -> config, utils, pdf_fonts
- Data Model Layer   : app.models     -> Pydantic schemas
- Database Layer     : app.db         -> SQLAlchemy (iso_size_ranges, iso_it_grades, iso_deviations)
- Auth & RBAC        : tại file main.py (JWT login, get_current_user, require_lecturer)

==> Chạy:  uvicorn main:app --reload
"""
import base64
import csv
import io
import json
import re
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from typing import List, Optional

import jwt
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from openai import OpenAI
from passlib.context import CryptContext
from pydantic import BaseModel
from sqlalchemy import Column, DateTime, Integer, String, Text
from sqlalchemy.orm import Session

from app.api.routes.export import router as export_router
from app.api.routes.iso import router as iso_router
from app.core.config import (
    API_TITLE,
    OPENROUTER_API_KEY,
    OPENROUTER_BASE_URL,
    OPENROUTER_VISION_MODEL,
)
from app.core.pdf_fonts import register_pdf_fonts
from app.db import seed as iso_seed
from app.db.database import Base, engine, get_db
from app.db import models as iso_models  # noqa: F401 (đảm bảo ORM models được đăng ký)
from app.db.models import User
from app.db.seed import seed_users_if_empty as seed_users


# ===========================================================================
# CẤU HÌNH JWT AUTH
# ===========================================================================
# ⚠️ Trong môi trường production hãy đặt biến môi trường JWT_SECRET_KEY
# để không dùng secret mặc định này.
import os

SECRET_KEY = os.getenv("JWT_SECRET_KEY", "iso286-edtech-secret-key-2026-change-me")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 8 * 60  # 8 giờ (1 buổi thi / học)


# ===========================================================================
# MOCK DATABASE - TÀI KHOẢN NGƯỜI DÙNG
# ===========================================================================
# ⚠️ Từ phiên bản này, toàn bộ tài khoản demo (giảng viên & sinh viên) KHÔNG còn
# được khai báo dạng dict trong bộ nhớ (in-memory) nữa.
#
# Mock data CHỈ chạy ngầm lúc STARTUP: `seed_users()` (app/db/seed.py) tự động
# tạo 3 GV (gv01..gv03) + 20 SV (2110481..2110500, mật khẩu '123') nếu bảng
# `users` rỗng. Auth (login) đọc trực tiếp từ CSDL SQL qua `db.query(User)`.
#
# => Không còn endpooint nào cho phép fetch danh sách user tự do ra ngoài.
#    Xem thêm hàm `_login_handler` & `_upload_students` phía dưới.
_pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def _hash_password(plain: str) -> str:
    """Băm mật khẩu bằng bcrypt (passlib). Dùng cho seed & upload sinh viên."""
    return _pwd_context.hash(plain)

# Role hiển thị tiếng Việt cho frontend
ROLE_LABELS = {
    "lecturer": "Giảng viên",
    "student": "Sinh viên",
}

# NOTE: Mọi tài khoản mẫu đều được seed vào CSDL ở startup (seed_users).
# Không khai báo USERS dict ở đây nữa để tránh dữ liệu mock "rò rỉ" ra giao diện.


# ===========================================================================
# MOCK ASSIGNMENTS - 3 ĐỀ THI MẪU (gán cho 3 Sinh viên đầu tiên)
# ===========================================================================
# Cấu trúc JSON tuân thủ đúng shape mà Frontend (index.html) cần:
#   buildExamDataFromBatch() / saveHistoryToBackend() dùng để "Xem lại" &
#   "Xuất Word/PDF". Mỗi phần tử gồm:
#   D, hole{sym,it,ES,EI,T,Dmax,Dmin}, shaft{sym,it,es,ei,T,dmax,dmin},
#   fit{type,class,colorTheme,Smax,Smin,Nmax,Nmin,TN},
#   student{mssv,name,className,classCode}, diffLabel, examCode, activeTasks[]
# Giá trị ES/EI es/ei (µm) và Dmax..dmin (mm) được tính theo chuẩn ISO 286
# (khớp với output của DualEngineMath.calculate trong index.html).
mock_assignments_db = [
    # ------------------------------------------------------------------
    # ĐỀ 1 - SV 2110481 : ϕ45 H7/g6  (Lắp lỏng - Clearance)
    # ------------------------------------------------------------------
    {
        "D": 45,
        "hole": {
            "sym": "H", "it": 7,
            "ES": 25, "EI": 0, "T": 25,
            "Dmax": 45.025, "Dmin": 45.000,
        },
        "shaft": {
            "sym": "g", "it": 6,
            "es": -9, "ei": -25, "T": 16,
            "dmax": 44.991, "dmin": 44.975,
        },
        "fit": {
            "type": "Lắp lỏng (Có độ hở)", "class": "clearance", "colorTheme": "emerald",
            "Smax": 50, "Smin": 9, "Nmax": -50, "Nmin": -9, "TN": 41,
        },
        "student": {
            "mssv": "2110481", "name": "Nguyễn Văn An",
            "className": "21CK1", "classCode": "ME2026",
        },
        "diffLabel": "Trung bình",
        "examCode": "ME2026-2110481",
        "activeTasks": [],
    },
    # ------------------------------------------------------------------
    # ĐỀ 2 - SV 2110482 : ϕ40 H7/k6  (Lắp trung gian - Transition)
    # ------------------------------------------------------------------
    {
        "D": 40,
        "hole": {
            "sym": "H", "it": 7,
            "ES": 25, "EI": 0, "T": 25,
            "Dmax": 40.025, "Dmin": 40.000,
        },
        "shaft": {
            "sym": "k", "it": 6,
            "es": 18, "ei": 2, "T": 16,
            "dmax": 40.018, "dmin": 40.002,
        },
        "fit": {
            "type": "Lắp trung gian", "class": "transition", "colorTheme": "amber",
            "Smax": 23, "Smin": -18, "Nmax": 18, "Nmin": -23, "TN": 41,
        },
        "student": {
            "mssv": "2110482", "name": "Trần Thị Bình",
            "className": "21CK1", "classCode": "ME2026",
        },
        "diffLabel": "Trung bình",
        "examCode": "ME2026-2110482",
        "activeTasks": [],
    },
    # ------------------------------------------------------------------
    # ĐỀ 3 - SV 2110483 : ϕ30 H7/p6  (Lắp chặt - Interference)
    # ------------------------------------------------------------------
    {
        "D": 30,
        "hole": {
            "sym": "H", "it": 7,
            "ES": 21, "EI": 0, "T": 21,
            "Dmax": 30.021, "Dmin": 30.000,
        },
        "shaft": {
            "sym": "p", "it": 6,
            "es": 35, "ei": 22, "T": 13,
            "dmax": 30.035, "dmin": 30.022,
        },
        "fit": {
            "type": "Lắp chặt (Có độ dôi)", "class": "interference", "colorTheme": "rose",
            "Smax": -35, "Smin": -1, "Nmax": 35, "Nmin": 1, "TN": 34,
        },
        "student": {
            "mssv": "2110483", "name": "Lê Văn Cường",
            "className": "21CK2", "classCode": "ME2026",
        },
        "diffLabel": "Trung bình",
        "examCode": "ME2026-2110483",
        "activeTasks": [],
    },
]


# ===========================================================================
# PYDANTIC SCHEMAS (Auth)
# ===========================================================================
class LoginRequest(BaseModel):
    username: str
    password: str


class UserInfo(BaseModel):
    username: str
    full_name: str
    role: str
    department: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserInfo


# ===========================================================================
# ORM MODEL - LƯU TRỮ LỊCH SỬ ĐỀ THI (generated_questions)
# ===========================================================================
class GeneratedQuestion(Base):
    """Bảng lưu lịch sử các đề thi đã tạo (mỗi hàng = 1 bài đề của 1 SV)."""

    __tablename__ = "generated_questions"

    id = Column(Integer, primary_key=True, autoincrement=True)
    batch_code = Column(String(32), nullable=False, index=True)
    student_mssv = Column(String(32), nullable=False)
    student_name = Column(String(128), nullable=False)
    exam_data = Column(Text, nullable=False, default="{}")  # JSON string chi tiết D, Lỗ, Trục
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)


# ===========================================================================
# PYDANTIC SCHEMAS (History)
# ===========================================================================
class HistoryItem(BaseModel):
    """Một bài đề trong payload POST /api/history."""

    student_mssv: str
    student_name: str
    exam_data: dict  # chi tiết D, Miền Lỗ, Miền Trục (JSON-serializable)


class HistoryCreateRequest(BaseModel):
    """Payload POST /api/history: lô đề + danh sách bài đề."""

    batch_code: str
    questions: List[HistoryItem]


class HistoryQuestionOut(BaseModel):
    """Một bài đề khi trả về cho frontend (GET /api/history)."""

    student_mssv: str
    student_name: str
    exam_data: dict


class HistoryBatchOut(BaseModel):
    """Một lô đề thi đã tạo (nhóm theo batch_code)."""

    batch_code: str
    created_at: datetime
    count: int
    questions: List[HistoryQuestionOut]


class HistorySaveResponse(BaseModel):
    """Response của POST /api/history: tóm tắt lô đề vừa lưu."""

    batch_code: str
    created_at: Optional[datetime] = None
    count: int


class StudentHistoryOut(BaseModel):
    """Một bài thi đã nộp của sinh viên (GET /api/history/student/{mssv}).

    Chỉ trả về các bài có trạng thái 'submitted' hoặc 'graded'.
    TUYỆT ĐỐI KHÔNG trả về bài đang ở trạng thái assigned (mới giao).
    """

    mssv: str
    examCode: str
    exam_data: dict          # Đề bài chi tiết: D, hole, shaft, fit, student...
    score: Optional[float] = None
    submitted_at: Optional[datetime] = None
    status: str


# ===========================================================================
# AUTHENTICATION HELPERS
# ===========================================================================
bearer_scheme = HTTPBearer(auto_error=False)


def create_access_token(username: str, role: str) -> str:
    """Tạo JWT token HS256 cho user (thời hạn cấu hình ở trên)."""
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    payload = {"sub": username, "exp": expire, "role": role}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(bearer_scheme),
    db: Session = Depends(get_db),
) -> UserInfo:
    """FastAPI dependency: xác thực JWT -> trả về user đang đăng nhập.

    Nếu token thiếu / sai / hết hạn -> HTTP 401.
    """
    if credentials is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Bạn cần đăng nhập để truy cập tài nguyên này.",
            headers={"WWW-Authenticate": "Bearer"},
        )
    try:
        payload = jwt.decode(
            credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM]
        )
        username = payload.get("sub")
        user = db.query(User).filter(User.username == username).first()
        if user is None:
            raise jwt.InvalidTokenError()
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Phiên đăng nhập đã hết hạn. Vui lòng đăng nhập lại.",
            headers={"WWW-Authenticate": "Bearer"},
        )
    except jwt.PyJWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token không hợp lệ hoặc đã bị chỉnh sửa.",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return UserInfo(**user.to_auth_dict())


def require_lecturer(user: UserInfo = Depends(get_current_user)) -> UserInfo:
    """FastAPI dependency: yêu cầu role = lecturer (chặn student).

    Dùng cho /api/export/docx và /api/export/pdf.
    """
    if user.role != "lecturer":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Chỉ Giảng viên (lecturer) mới được sử dụng tính năng xuất file Word/PDF.",
        )
    return user


# ===========================================================================
# LOGIN ENDPOINT
# ===========================================================================
def _login_handler(payload: LoginRequest, db: Session = Depends(get_db)) -> LoginResponse:
    """Xử lý đăng nhập: kiểm tra user + mật khẩu trong CSDL -> cấp JWT."""
    user = db.query(User).filter(User.username == payload.username).first()
    if user is None or not _pwd_context.verify(payload.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Sai tên đăng nhập hoặc mật khẩu.",
        )
    token = create_access_token(user.username, user.role)
    return LoginResponse(
        access_token=token,
        user=UserInfo(**user.to_auth_dict()),
    )


# ===========================================================================
# PYDANTIC SCHEMAS (Assignments - Bài tập Sinh viên)
# ===========================================================================
class AssignmentSubmitRequest(BaseModel):
    """Payload POST /api/assignments/{mssv}/submit: câu trả lời của sinh viên.

    Gồm 4 sai lệch giới hạn (µm):
      - ES : Sai lệch giới hạn trên của Lỗ (Hole, Upper Deviation)
      - EI : Sai lệch giới hạn dưới của Lỗ (Hole, Lower Deviation)
      - es : Sai lệch giới hạn trên của Trục (Shaft, Upper Deviation)
      - ei : Sai lệch giới hạn dưới của Trục (Shaft, Lower Deviation)
    """

    ES: float
    EI: float
    es: float
    ei: float


# ===========================================================================
# ASSIGNMENT HELPERS (Tìm & Chấm điểm bài tập)
# ===========================================================================
def _find_assignment(mssv: str) -> dict:
    """Tìm đề thi của sinh viên trong mock_assignments_db.

    Nếu không tồn tại đề -> HTTP 404 (Idiomatic REST).

    ⚠️ Chỉ dùng cho các endpoint cần đề thi BẮT BUỘC (VD: nộp bài).
    Đối với endpoint GET lấy bài tập, hãy dùng `_find_assignment_or_none`
    để trả về trạng thái rỗng (Empty State) thân thiện thay vì 404.
    """
    # Chuẩn hoá mssv: kiểu số (2110481) hoặc chữ ("2110481") đều hợp lệ.
    mssv = str(mssv).strip()
    for assignment in mock_assignments_db:
        student = assignment.get("student", {})
        if str(student.get("mssv", "")).strip() == mssv:
            return assignment
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"Không tìm thấy đề thi cho MSSV {mssv}. Liên hệ Giảng viên để được cấp đề.",
    )


def _find_assignment_or_none(mssv: str):
    """Tìm đề thi nhưng KHÔNG raise 404 khi không tìm thấy.

    Trả về:
      - assignment dict nếu tìm thấy.
      - None nếu sinh viên chưa được giao bài.
    """
    mssv = str(mssv).strip()
    for assignment in mock_assignments_db:
        student = assignment.get("student", {})
        if str(student.get("mssv", "")).strip() == mssv:
            return assignment
    return None


def _get_assignment_state(mssv: str) -> dict:
    """Lấy state (status/score/student_answers/correct key) của một đề.

    Mặc định chưa nộp: status = 'not_submitted', score = None.
    """
    assignment = _find_assignment(mssv)
    hole, shaft = assignment["hole"], assignment["shaft"]
    correct_answers = {
        "ES": hole["ES"],
        "EI": hole["EI"],
        "es": shaft["es"],
        "ei": shaft["ei"],
    }
    return {
        "mssv": mssv,
        "assignment": {
            "D": assignment["D"],
            "hole": hole,
            "shaft": shaft,
            "fit": assignment.get("fit", {}),
            "student": assignment.get("student", {}),
            "diffLabel": assignment.get("diffLabel", ""),
            "examCode": assignment.get("examCode", ""),
        },
        "status": assignment.get("status", "not_submitted"),
        "score": assignment.get("score"),
        "student_answers": assignment.get("student_answers"),
        "correct_answers": correct_answers,
    }


def _grade_assignment(mssv: str, answers: AssignmentSubmitRequest) -> dict:
    """Chấm điểm tự động (Auto-grading) bài tập của sinh viên.

    Quy tắc chấm:
      - Mỗi thông số đúng (ép kiểu float trước khi so sánh) được +2.5 điểm.
      - 4 thông số (ES, EI, es, ei) -> Tối đa 10.0 điểm.
      - Điểm số được làm tròn 1 chữ số thập phân để tránh sai số dấu phẩy động.

    Nếu đề đã nộp trước đó -> trả về kết quả đã lưu (không chấm lại).
    """
    state = _get_assignment_state(mssv)

    # ❌ Đã nộp rồi: trả lại kết quả cũ, không cho nộp lại.
    if state["status"] == "submitted" and state["score"] is not None:
        return {
            "mssv": mssv,
            "status": "submitted",
            "already_submitted": True,
            "score": state["score"],
            "correct_answers": state["correct_answers"],
            "student_answers": state["student_answers"] or {},
            "details": _build_result_details(state["correct_answers"], state["student_answers"] or {}),
        }

    assignment = _find_assignment(mssv)
    # Ép kiểu dữ liệu float trước khi so sánh để tránh lỗi (string vs number).
    student_answers = {
        "ES": float(answers.ES),
        "EI": float(answers.EI),
        "es": float(answers.es),
        "ei": float(answers.ei),
    }

    correct_answers = state["correct_answers"]
    result_details = _build_result_details(correct_answers, student_answers)

    score = round(sum(
        2.5 for item in result_details if item["is_correct"]
    ), 1)

# ✅ Cập nhật trạng thái đề thi trong DB (mock_assignments_db).
    assignment["status"] = "submitted"
    assignment["score"] = score
    assignment["student_answers"] = student_answers
    assignment["submitted_at"] = datetime.utcnow()  # Thời điểm nộp bài (cho lịch sử)

    return {
        "mssv": mssv,
        "status": "submitted",
        "already_submitted": False,
        "score": score,
        "correct_answers": correct_answers,
        "student_answers": student_answers,
        "details": result_details,
    }


def _build_result_details(correct_answers: dict, student_answers: dict) -> list:
    """Tạo bảng đối chiếu từng thông số: (key, correct, student, is_correct)."""
    labels = {
        "ES": "Sai lệch trên Lỗ - ES (µm)",
        "EI": "Sai lệch dưới Lỗ - EI (µm)",
        "es": "Sai lệch trên Trục - es (µm)",
        "ei": "Sai lệch dưới Trục - ei (µm)",
    }
    details = []
    for key in ("ES", "EI", "es", "ei"):
        correct_val = float(correct_answers[key])
        student_val = float(student_answers.get(key, 0))
        details.append({
            "key": key,
            "label": labels[key],
            "correct": correct_answers[key],
            "student": student_val,
            "is_correct": abs(correct_val - student_val) < 1e-9,
        })
    return details


# ===========================================================================
# ASSIGNMENT ENDPOINTS (SINH VIÊN LÀM & NỘP BÀI - AUTO-GRADING)
# ===========================================================================
def _get_assignment(mssv: str) -> dict:
    """GET /api/assignments/{mssv}: trả về đề thi + trạng thái nộp bài.

    Frontend dùng trường `status` để quyết định:
      - 'not_submitted' -> hiện Form nhập liệu.
      - 'submitted'     -> ẩn Form, hiện Thẻ Kết Quả (Result Card).
      - 'empty'         -> sinh viên chưa được giao bài (Empty State).

    Graceful Degradation: nếu KHÔNG tìm thấy đề thi của mssv, KHÔNG raise 404.
    Thay vào đó trả về HTTP 200 kèm JSON báo hiệu rỗng:
      {"status": "empty", "message": "Bạn chưa có bài tập mới nào."}
    """
    assignment = _find_assignment_or_none(mssv)
    if assignment is None:
        # Graceful Degradation: KHÔNG dùng raise HTTPException(404).
        # Trả về HTTP 200 + status "empty" để frontend hiển thị Empty State thân thiện.
        return {
            "status": "empty",
            "message": "Bạn chưa có bài tập mới nào.",
            "mssv": str(mssv),
        }
    return _get_assignment_state(mssv)


def _submit_assignment(mssv: str, payload: AssignmentSubmitRequest) -> dict:
    """POST /api/assignments/{mssv}/submit: nhận câu trả lời -> chấm & lưu điểm.

    Payload JSON: {"ES": ..., "EI": ..., "es": ..., "ei": ...}
    """
    return _grade_assignment(mssv, payload)


# ===========================================================================
# HISTORY ENDPOINTS (LƯU TRỮ LỊCH SỬ ĐỀ THI)
# ===========================================================================
def _save_history(payload: HistoryCreateRequest, db: Session = Depends(get_db)) -> dict:
    """POST /api/history: bulk insert lô đề thi vào bảng generated_questions.

    Nhận payload từ Frontend sau khi tạo đề xong và lưu hàng loạt.
    """
    if not payload.batch_code.strip():
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="batch_code không được để trống.",
        )
    if not payload.questions:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Lô đề không chứa bài thi nào.",
        )

    rows = [
        GeneratedQuestion(
            batch_code=payload.batch_code,
            student_mssv=q.student_mssv,
            student_name=q.student_name,
            exam_data=json.dumps(q.exam_data, ensure_ascii=False),
        )
        for q in payload.questions
    ]

    db.add_all(rows)
    db.commit()

    # Trả về batch vừa lưu (để frontend có thể refresh hoặc hiển thị ngay).
    return {
        "batch_code": payload.batch_code,
        "created_at": rows[0].created_at,
        "count": len(rows),
    }


def _list_history(db: Session = Depends(get_db)) -> List[HistoryBatchOut]:
    """GET /api/history: trả về danh sách lô đề đã tạo (nhóm theo batch_code).

    Sắp xếp lô mới nhất lên đầu. Mỗi lô trả kèm:
    - batch_code, created_at (thời điểm tạo lô đầu tiên)
    - count (số bài thi trong lô)
    - questions (chi tiết từng bài: mssv, tên, exam_data JSON)
    """
    # Lấy tất cả bản ghi, sắp theo thời gian tạo mới nhất.
    rows = (
        db.query(GeneratedQuestion)
        .order_by(GeneratedQuestion.created_at.desc(), GeneratedQuestion.id.desc())
        .all()
    )

    # Nhóm theo batch_code (giữ thứ tự mới nhất).
    batches: dict = {}
    for row in rows:
        if row.batch_code not in batches:
            batches[row.batch_code] = {
                "batch_code": row.batch_code,
                "created_at": row.created_at,
                "questions": [],
            }
        try:
            exam_data = json.loads(row.exam_data) if row.exam_data else {}
        except (json.JSONDecodeError, TypeError):
            exam_data = {}
        batches[row.batch_code]["questions"].append(
            {
                "student_mssv": row.student_mssv,
                "student_name": row.student_name,
                "exam_data": exam_data,
            }
        )

    result = []
    for code, batch in batches.items():
        result.append(
            HistoryBatchOut(
                batch_code=batch["batch_code"],
                created_at=batch["created_at"],
                count=len(batch["questions"]),
                questions=batch["questions"],
            )
        )
    return result


def _list_student_history(mssv: str) -> List[StudentHistoryOut]:
    """GET /api/history/student/{mssv}: lịch sử bài đã NỘP của một sinh viên.

    QUYỀN SINH VIÊN:
    - Chỉ trả về các bài tập THUỘC VỀ đúng mssv đó.
    - Chỉ lấy những bài có status ∈ {"submitted", "graded"} (đã nộp / đã chấm).
    - TUYỆT ĐỐI KHÔNG trả về bài đang ở trạng thái 'assigned' (mới được giao,
      chưa làm) hoặc 'not_submitted' (chưa nộp).

    Nếu mssv không tồn tại hoặc chưa nộp bài nào -> trả về danh sách rỗng ([]),
    để frontend hiển thị Empty State thân thiện.
    """
    mssv = str(mssv).strip()
    allowed_statuses = {"submitted", "graded"}
    result: List[StudentHistoryOut] = []

    for assignment in mock_assignments_db:
        student = assignment.get("student", {})
        if str(student.get("mssv", "")).strip() != mssv:
            continue  # (1) Không thuộc về sinh viên này -> bỏ qua.
        status = assignment.get("status", "not_submitted")
        if status not in allowed_statuses:
            continue  # (2) Chưa nộp / mới giao -> TUYỆT ĐỐI bỏ qua.

        # Đề bài chi tiết (D, Lỗ, Trục, đặc tính lắp ghép...).
        exam_data = {
            "D": assignment.get("D"),
            "hole": assignment.get("hole", {}),
            "shaft": assignment.get("shaft", {}),
            "fit": assignment.get("fit", {}),
            "student": student,
            "diffLabel": assignment.get("diffLabel", ""),
            "examCode": assignment.get("examCode", ""),
        }

        result.append(
            StudentHistoryOut(
                mssv=mssv,
                examCode=assignment.get("examCode", ""),
                exam_data=exam_data,
                score=assignment.get("score"),
                submitted_at=assignment.get("submitted_at"),
                status=status,
            )
        )

    # Sắp xếp bài mới nhất (theo thời gian nộp) lên đầu.
    result.sort(key=lambda r: r.submitted_at or datetime.min, reverse=True)
    return result


# ===========================================================================
# AI OCR CHẤM BÀI TỰ ĐỘNG (OpenRouter Vision Model)
# ===========================================================================
# Hệ thống dùng OpenRouter (chuẩn API OpenAI) với model Vision miễn phí
# `google/gemini-1.5-flash` để đọc ảnh bài làm và trích xuất các sai lệch
# giới hạn (ES, EI, es, ei, Smax, Smin, Nmax, Nmin) theo tiêu chuẩn ISO 286.

# System Prompt gửi kèm ảnh -> model trả đúng định dạng JSON.
_AI_SYSTEM_PROMPT = (
    "Đây là bài làm môn dung sai lắp ghép ISO 286. Hãy đọc ảnh và trích xuất "
    "các thông số sinh viên đã giải: ES, EI, es, ei, Smax, Smin, Nmax, Nmin. "
    'Trả về đúng định dạng JSON, ví dụ: {"ES": 35, "EI": 0}. Không giải thích thêm.'
)

# Client OpenAI tương thích OpenRouter (base_url riêng).
# ⚠️ Khởi tạo LAZY (chỉ tạo khi thật sự gọi /api/ai-grade) để ứng dụng vẫn
#    chạy bình thường kể cả khi chưa cấu hình OPENROUTER_API_KEY.
_openai_client = None


def _get_openai_client():
    """Trả về OpenAI client (OpenRouter). Tạo lazy lần đầu khi được gọi.:
    """
    global _openai_client
    if _openai_client is None:
        _openai_client = OpenAI(
            api_key=OPENROUTER_API_KEY,
            base_url=OPENROUTER_BASE_URL,
        )
    return _openai_client


def _extract_json_from_llm(text: str) -> dict:
    """Trích xuất JSON dict từ phản hồi văn bản của model.

    Model đôi khi bọc JSON trong ```json ... ``` (markdown fenced code block)
    hoặc thêm chữ giải thích. Hàm này:
      1. Bỏ thẻ ```json ... ``` nếu có.
      2. Tìm khối `{...}` đầu tiên trong text.
      3. Parse bằng json.loads.

    Nếu không tìm thấy/parse lỗi -> raise ValueError (để handler bắt lỗi).
    """
    # Bỏ toàn bộ thẻ markdown code fence (```json ... ```).
    cleaned = re.sub(r"```(?:json)?", "", text).strip()
    # Tìm khối JSON đầu tiên (từ `{` đến `}`).
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("Model không trả về JSON hợp lệ.")

    raw = cleaned[start : end + 1]
    data = json.loads(raw)
    if not isinstance(data, dict):
        raise ValueError("Kết quả model không phải là một JSON object.")
    return data


def _ai_grade_handler(file: UploadFile):
    """POST /api/ai-grade: nhận ảnh bài làm -> OCR bằng AI -> trả JSON thông số.

    Quy trình:
      1. Đọc bytes của file ảnh tải lên.
      2. Chuyển đổi sang Base64 (data URL: data:image/<ext>;base64,...).
      3. Gửi lên OpenRouter (model vision google/gemini-1.5-flash) kèm
         System Prompt yêu cầu trích xuất ES, EI, es, ei, Smax, Smin, Nmax, Nmin.
      4. Trích xuất JSON từ phản hồi và trả về frontend.

    Lỗi (không đọc được file, API lỗi, JSON không hợp lệ) -> HTTP 400/502.
    """
    try:
        # (1) Đọc toàn bộ nội dung file ảnh (bytes).
        content = file.file.read()
        if not content:
            raise ValueError("File ảnh rỗng hoặc không đọc được.")

        # (2) Xác định MIME type từ extension (mặc định image/png).
        ext = file.filename.rsplit(".", 1)[-1].lower() if file.filename else "png"
        mime = f"image/{ext}" if ext in {"png", "jpg", "jpeg", "gif", "webp"} else "image/png"

        # (2b) Chuyển ảnh sang Base64 dạng data URL (chuẩn cho Vision API).
        b64_image = base64.b64encode(content).decode("utf-8")
        data_url = f"data:{mime};base64,{b64_image}"

# (3) Gọi model vision trên OpenRouter.
        client = _get_openai_client()
        response = client.chat.completions.create(
            model=OPENROUTER_VISION_MODEL,
            messages=[
                {"role": "system", "content": _AI_SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "Hãy đọc bài làm trong ảnh và trả về JSON các thông số.",
                        },
                        {"type": "image_url", "image_url": {"url": data_url}},
                    ],
                },
            ],
            response_format={"type": "json_object"},
            max_tokens=1024,
        )

        # (4) Lấy nội dung trả về và trích xuất JSON.
        llm_text = response.choices[0].message.content or ""
        result = _extract_json_from_llm(llm_text)

        return {
            "success": True,
            "data": result,
            "model": OPENROUTER_VISION_MODEL,
        }
    except ValueError as exc:
        # Lỗi nghiệp vụ (file rỗng, JSON không hợp lệ).
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    except Exception as exc:
        # Lỗi hạ tầng (OpenRouter timeout, API key sai, network...).
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Không gọi được AI OCR: {exc}",
        )


# ===========================================================================
# UPLOAD DANH SÁCH SINH VIÊN (POST /api/students/upload)
# ===========================================================================
# Nhận file .csv / .xlsx (cột: MSSV, HoTen, Lop) -> bulk insert vào bảng users
# với role='student', mật khẩu mặc định '123' (hash bcrypt).
# Chỉ Giảng viên (lecturer) mới được gọi.
def _parse_student_file(file: UploadFile):
    """Đọc file tải lên (.csv/.xlsx) và trích xuất danh sách {mssv, name, lop}."""
    filename = (file.filename or "").lower()
    if not filename.endswith(".csv") and not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Định dạng không hợp lệ. Chỉ hỗ trợ file .csv hoặc .xlsx.",
        )

    content = file.file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File rỗng hoặc không đọc được.",
        )

    students = []
    if filename.endswith(".csv"):
        # Đọc CSV (hỗ trợ nhiều encoding để tránh lỗi tiếng Việt).
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            mssv = (row.get("MSSV") or row.get("mssv") or "").strip()
            name = (row.get("HoTen") or row.get("hoten") or row.get("Họ Tên") or "").strip()
            lop = (row.get("Lop") or row.get("lop") or row.get("Lớp") or "").strip()
            if mssv:
                students.append({"mssv": mssv, "name": name, "lop": lop})
    else:
        # Đọc XLSX bằng openpyxl.
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Thiếu thư viện openpyxl. Vui lòng cài: pip install openpyxl",
            )
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File Excel không có dữ liệu (thiếu header).",
            )
        header = [str(h).strip().lower() if h else "" for h in header]
        idx_mssv = next((i for i, h in enumerate(header) if h in ("mssv", "mã số sv", "mã sinh viên")), 0)
        idx_name = next((i for i, h in enumerate(header) if h in ("hoten", "họ và tên", "họ tên", "tên")), 1)
        idx_lop = next((i for i, h in enumerate(header) if h in ("lop", "lớp")), 2)
        for row in rows_iter:
            if not row:
                continue
            mssv = str(row[idx_mssv]).strip() if idx_mssv < len(row) and row[idx_mssv] is not None else ""
            name = str(row[idx_name]).strip() if idx_name < len(row) and row[idx_name] is not None else ""
            lop = str(row[idx_lop]).strip() if idx_lop < len(row) and row[idx_lop] is not None else ""
            if mssv:
                students.append({"mssv": mssv, "name": name, "lop": lop})

    return students


def _upload_students(file: UploadFile, db: Session = Depends(get_db)) -> dict:
    """POST /api/students/upload: đọc file -> bulk insert sinh viên vào bảng users."""
    students = _parse_student_file(file)
    if not students:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Không tìm thấy dữ liệu sinh viên hợp lệ trong file (cột MSSV, HoTen, Lop).",
        )

    created = 0
    skipped = 0
    for s in students:
        exists = db.query(User).filter(User.username == s["mssv"]).first()
        if exists:
            skipped += 1
            continue
        db.add(User(
            username=s["mssv"],
            full_name=s["name"] or f"Sinh viên {s['mssv']}",
            role="student",
            department=s["lop"],
            password_hash=_pwd_context.hash("123"),
        ))
        created += 1

    db.commit()
    return {
        "success": True,
        "created": created,
        "skipped": skipped,
        "total": len(students),
        "message": f"Đã thêm {created} sinh viên, bỏ qua {skipped} tài khoản đã tồn tại.",
    }


def _list_students(db: Session = Depends(get_db)) -> dict:
    """GET /api/students: trả về danh sách sinh viên (role='student').

    Dùng để Frontend cập nhật bảng "Danh sách sinh viên" sau khi upload.
    """
    students = (
        db.query(User)
        .filter(User.role == "student")
        .order_by(User.username)
        .all()
    )
    return {
        "success": True,
        "total": len(students),
        "students": [u.to_dict() for u in students],
    }


# ===========================================================================
# UPLOAD DANH SÁCH SINH VIÊN - ENDPOINT GIẢNG VIÊN (POST /api/teachers/upload-students)
# ===========================================================================
# 📌 Endpoint mới theo yêu cầu: dùng pandas để đọc nội dung file .csv/.xlsx
#   trực tiếp trong memory (không cần typeof openpyxl thủ công), trích xuất
#   3 cột MSSV, HoTen, Lop, rồi bulk insert vào bảng users với role='student'
#   và mật khẩu mặc định '123456'.
#
#   Khác biệt so với endpoint cũ /api/students/upload:
#     - Dùng pandas (hỗ trợ cả .csv & .xlsx cùng 1 code path).
#     - Mật khẩu mặc định là '123456' (theo yêu cầu).
#     - Route mới: /api/teachers/upload-students.
#   Cả hai endpoint đều chỉ Giảng viên (lecturer) mới được gọi.
def _parse_student_file_pandas(file: UploadFile):
    """Đọc file .csv/.xlsx bằng pandas và trích xuất danh sách {mssv, name, lop}.

    Cột dữ liệu cần có: MSSV, HoTen, Lop.
    Hàm cũng tự nhận diện các tên cột linh hoạt:
      - MSSV: 'MSSV', 'Mã số SV', 'MaSV', 'student_id', ...
      - HoTen: 'HoTen', 'Họ và tên', 'Họ Tên', 'full_name', ...
      - Lop: 'Lop', 'Lớp', 'class', ...
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".csv") and not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Định dạng không hợp lệ. Chỉ hỗ trợ file .csv hoặc .xlsx.",
        )

    content = file.file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File rỗng hoặc không đọc được.",
        )

    # Kiểm tra pandas đã được cài hay chưa (thêm vào requirements.txt).
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Thiếu thư viện pandas. Vui lòng cài: pip install pandas openpyxl",
        )

    try:
        if filename.endswith(".csv"):
            # pandas tự động đoán encoding (thường UTF-8 BOM / UTF-8 / Latin-1).
            df = pd.read_csv(io.BytesIO(content))
        else:
            # pandas đọc .xlsx trực tiếp từ memory (BytesIO) — không ghi file tạm.
            df = pd.read_excel(io.BytesIO(content))
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Không đọc được nội dung file: {exc}",
        )

    # Chuẩn hoá tên cột: bỏ dấu cách thừa, ký tự đặc biệt, viết thường.
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Nhận diện tên cột linh hoạt.
    def _find_col(*candidates):
        for c in candidates:
            for col in df.columns:
                if col == c or c in col:
                    return col
        return None

    col_mssv = _find_col("mssv", "mã số sv", "masv", "student_id", "msv")
    col_name = _find_col("hoten", "họ và tên", "họ tên", "full_name", "tên", "ten")
    col_lop = _find_col("lop", "lớp", "class", "classname")

    if col_mssv is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Không tìm thấy cột 'MSSV' trong file. Cần ít nhất cột MSSV.",
        )

    students = []
    for _, row in df.iterrows():
        mssv = str(row[col_mssv]).strip()
        if not mssv or mssv.lower() in ("nan", "none", "null"):
            continue
        name = str(row[col_name]).strip() if col_name else ""
        lop = str(row[col_lop]).strip() if col_lop else ""
        # Loại bỏ chuỗi 'nan' do pandas trả về cho ô trống.
        if name.lower() in ("nan", "none", "null"):
            name = ""
        if lop.lower() in ("nan", "none", "null"):
            lop = ""
        students.append({"mssv": mssv, "name": name, "lop": lop})

    return students


def _upload_students_teachers(file: UploadFile, db: Session = Depends(get_db)) -> dict:
    """POST /api/teachers/upload-students: nhận file -> bulk insert sinh viên.

    - Đọc file .csv/.xlsx bằng pandas (trong memory).
    - Trích xuất 3 cột: MSSV, HoTen, Lop.
    - Bulk insert vào bảng users với role='student', mật khẩu mặc định '123456'.
    - Bỏ qua những sinh viên có MSSV đã tồn tại (tránh lỗi duplicate khóa chính).
    - Trả về báo cáo: "Đã thêm thành công X sinh viên, bỏ qua Y sinh viên trùng lặp".
    """
    students = _parse_student_file_pandas(file)
    if not students:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Không tìm thấy dữ liệu sinh viên hợp lệ trong file (cần cột MSSV, HoTen, Lop).",
        )

    created = 0
    skipped = 0
    for s in students:
        exists = db.query(User).filter(User.username == s["mssv"]).first()
        if exists:
            skipped += 1
            continue
        db.add(User(
            username=s["mssv"],
            full_name=s["name"] or f"Sinh viên {s['mssv']}",
            role="student",
            department=s["lop"],
            password_hash=_pwd_context.hash("123456"),
        ))
        created += 1

    db.commit()
    return {
        "success": True,
        "created": created,
        "skipped": skipped,
        "total": len(students),
        "message": f"Đã thêm thành công {created} sinh viên, bỏ qua {skipped} sinh viên trùng lặp.",
    }


# ===========================================================================
# LIFESPAN (KHỞI TẠO DB + SEED)
# ===========================================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Khởi tạo CSDL + seed dữ liệu ISO 286 khi ứng dụng khởi động."""
    Base.metadata.create_all(bind=engine)
    seeded = iso_seed.seed_if_empty()
    # Tự động tạo tài khoản mẫu (3 GV + 20 SV, mật khẩu '123') nếu DB trống.
    users_created = seed_users()
    print(f"[ISO 286 DB] Bảng tra đã sẵn sàng. Seed mới: {seeded}")
    print(f"[AUTH] Đã tạo {users_created} tài khoản mẫu (gv01..gv03, 2110481..2110500, mật khẩu: 123)")
    yield


app = FastAPI(title=API_TITLE, lifespan=lifespan)

# Đăng ký font PDF (Times New Roman Unicode) một lần khi khởi động.
register_pdf_fonts()

# CORS: cho phép mọi origin để nút "Xuất Word/PDF" hoạt động
# dù trang được mở từ Live Server (5500), localhost (3000) hay file://
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# TRANG CHỦ (SERVE FRONTEND index.html)
# ---------------------------------------------------------------------------
# Để app có thể chạy trên một URL duy nhất (VD: Render), ta phục vụ luôn
# file frontend index.html tại đường dẫn gốc "/". Khi deploy trên Render,
# mọi người chỉ cần truy cập URL app là thấy giao diện (không cần Live Server).
import os as _os
_INDEX_HTML = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "index.html")


@app.get("/", include_in_schema=False)
async def _serve_index():
    return FileResponse(_INDEX_HTML)


# ---------------------------------------------------------------------------
# ĐĂNG KÝ ROUTER
# ---------------------------------------------------------------------------
# /api/login - công khai, không cần token
app.add_api_route(
    "/api/login",
    _login_handler,
    methods=["POST"],
    response_model=LoginResponse,
    tags=["auth"],
    summary="Đăng nhập hệ thống (cấp JWT token)",
)

# /api/iso/* - MỌI user đã đăng nhập (lecturer & student) đều truy cập được.
#   -> bảo vệ bằng Depends(get_current_user)
app.include_router(
    iso_router,
    prefix="/api",
    dependencies=[Depends(get_current_user)],
)

# /api/export/* - CHỈ lecturer mới được gọi.
#   -> bảo vệ bằng Depends(require_lecturer)
app.include_router(
    export_router,
    prefix="/api",
    dependencies=[Depends(require_lecturer)],
)

# /api/history - MỌI user đã đăng nhập (lecturer & student) đều truy cập được.
#   -> bảo vệ bằng Depends(get_current_user)
app.add_api_route(
    "/api/history",
    _save_history,
    methods=["POST"],
    response_model=HistorySaveResponse,
    dependencies=[Depends(get_current_user)],
    tags=["history"],
    summary="Lưu lô đề thi đã tạo vào cơ sở dữ liệu (bulk insert)",
)
app.add_api_route(
    "/api/history",
    _list_history,
    methods=["GET"],
    response_model=List[HistoryBatchOut],
    dependencies=[Depends(get_current_user)],
    tags=["history"],
    summary="Lấy danh sách lô đề thi đã tạo (nhóm theo batch_code)",
)

# /api/history/student/{mssv} - Lịch sử bài đã NỘP của một sinh viên.
#   -> chỉ trả về bài có status ∈ {"submitted", "graded"} (đã nộp / đã chấm).
#   -> TUYỆT ĐỐI KHÔNG trả về bài đang ở trạng thái assigned (mới giao, chưa làm).
#   -> bảo vệ bằng Depends(get_current_user)
app.add_api_route(
    "/api/history/student/{mssv}",
    _list_student_history,
    methods=["GET"],
    response_model=List[StudentHistoryOut],
    dependencies=[Depends(get_current_user)],
    tags=["history"],
    summary="Lấy lịch sử bài đã NỘP của sinh viên (chỉ trả về bài đã nộp / đã chấm)",
)

# /api/assignments/* - MỌI user đã đăng nhập (lecturer & student) đều truy cập được.
#   -> bảo vệ bằng Depends(get_current_user)
# GET  /api/assignments/{mssv}           : lấy đề thi + trạng thái nộp bài.
# POST /api/assignments/{mssv}/submit    : nộp bài & chấm điểm tự động (auto-grading).
app.add_api_route(
    "/api/assignments/{mssv}",
    _get_assignment,
    methods=["GET"],
    dependencies=[Depends(get_current_user)],
    tags=["assignments"],
    summary="Lấy đề thi của sinh viên + trạng thái nộp bài",
)
app.add_api_route(
    "/api/assignments/{mssv}/submit",
    _submit_assignment,
    methods=["POST"],
    dependencies=[Depends(get_current_user)],
    tags=["assignments"],
    summary="Nộp bài tập & chấm điểm tự động (Auto-grading)",
)

# /api/ai-grade - MỌI user đã đăng nhập (lecturer & student) đều truy cập được.
#   -> bảo vệ bằng Depends(get_current_user)
#   -> nhận ảnh bài làm (UploadFile) -> OCR bằng AI (OpenRouter vision model)
#      -> trả JSON các sai lệch giới hạn (ES, EI, es, ei, Smax, Smin, Nmax, Nmin).
app.add_api_route(
    "/api/ai-grade",
    _ai_grade_handler,
    methods=["POST"],
    dependencies=[Depends(get_current_user)],
    tags=["ai-grade"],
    summary="AI OCR chấm bài: đọc ảnh bài làm và trích xuất thông số ISO 286",
)

# /api/students/upload - CHỈ lecturer mới được gọi.
#   -> nhận file .csv/.xlsx -> bulk insert sinh viên vào bảng users (role=student, pass '123').
app.add_api_route(
    "/api/students/upload",
    _upload_students,
    methods=["POST"],
    dependencies=[Depends(require_lecturer)],
    tags=["students"],
    summary="Upload danh sách sinh viên (.csv/.xlsx) và tạo tài khoản hàng loạt",
)

# /api/students - CHỈ lecturer mới được gọi.
#   -> trả về danh sách sinh viên (role='student') để Frontend cập nhật bảng sau khi upload.
app.add_api_route(
    "/api/students",
    _list_students,
    methods=["GET"],
    dependencies=[Depends(require_lecturer)],
    tags=["students"],
    summary="Lấy danh sách sinh viên (role=student)",
)

# /api/teachers/upload-students - CHỈ lecturer mới được gọi.
#   -> nhận file .csv/.xlsx -> pandas đọc trong memory -> bulk insert sinh viên
#      vào bảng users (role=student, mật khẩu mặc định '123456').
#   -> bỏ qua MSSV đã tồn tại (thoát lỗi duplicate khóa chính).
app.add_api_route(
    "/api/teachers/upload-students",
    _upload_students_teachers,
    methods=["POST"],
    dependencies=[Depends(require_lecturer)],
    tags=["students"],
    summary="Upload danh sách sinh viên (pandas, pass '123456') từ màn hình Giảng viên",
)

